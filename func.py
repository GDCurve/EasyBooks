from openpyxl import load_workbook
import sys
import subprocess

def load():
    Book = load_workbook("Data.xlsx")
    Sheet = Book['Sheet']
    return Book, Sheet


def startScreen():

    print("""Welcome, please choose an option:\n
        edit | count | product | help | exit | gui """)
    Choice = input('Choice >>> ').lower()

    def edit_choice():
        clearscreen()
        id = input('Input ID of product >>> ')
        if id.isdigit():
            edit(id)
        else:
            print('Invalid ID')
            startScreen()

    def count_choice():
        clearscreen()
        id = input('Input ID of product >>> ')
        if id.isdigit():
            count(id)
        else:
            print('Invalid ID')
            startScreen()

    def product_choice():
        clearscreen()
        product()

    def help_choice():
        clearscreen()
        print(r"""--------------------------- HELP -----------------------------
edit => allows for addition or subtraction of product count.

count => reads the product count in warehouse

product => allows adding, removing or listing products

help => opens help menu

gui => opens the GUI mode (WIP)

exit => exits the application
---------------------------------------------------------------""")
        startScreen()

    def exit_choice():
        print('exiting...')
        sys.exit()

    def default_choice():
        print('unknown command, try again')
        startScreen()

    switcher = {
        "edit": edit_choice,
        "count": count_choice,
        "product": product_choice,
        "help": help_choice,
        "exit": exit_choice,
        "gui": start_gui
    }

    func = switcher.get(Choice, default_choice)

    func()

def edit(id):
    Book,Sheet = load()

    name, qty, i = finder(id)

    if finder(id) == (None, None, None):
        print("Product not found")
        startScreen()

    clearscreen()
    print('You are about to edit the count of ' + name + " which is stored at a quantity of " + str(qty))
    ans = input('Continue? Y / N >>> ').lower()
    if ans == "y":
        qty = input('Edit count by how much? >>> ')

        try:
            int(qty)
        except ValueError:
            print('Invalid quantity')
            startScreen()

        Sheet.cell(row=i, column=3).value = Sheet.cell(row=i, column=3).value + int(qty)
        Book.save('Data.xlsx')
        input('New count of ' + name + ' is ' + str(Sheet.cell(row=i, column=3).value) +
               ' press enter to continue')
        startScreen()
    else:
        startScreen()


def count(id):
    Book,Sheet = load()
    name, qty, i = finder(id)

    if finder(id) == (None, None, None):
        print("Product not found")
        startScreen()

    print("there's " + str(qty) + " of " + str(name) + " stored")

    ans = input('Count another product? Y/N >>> ').lower()
    if ans == "y":
        id = input('Input ID of product >>> ')
        count(id)
    elif ans == "n":
        startScreen()
    else:
        print('Unknown command')


def product():
    print(r""" choose an option:
        list | add | remove""")
    choice = input("Choice >>> ").lower()

    def list_choice():
        clearscreen()
        product_list()

    def add_choice():
        clearscreen()
        product_add()

    def remove_choice():
        clearscreen()
        id = input('Input ID of product >>> ')
        if id.isdigit():
            product_remove(id)
        else:
            print('Invalid ID')
            startScreen()


    def default_choice():
        print("unknown choice, try again")
        product()

    switcher = {
        "list": list_choice,
        "add": add_choice,
        "remove": remove_choice,
    }

    func = switcher.get(choice, default_choice)

    func()


# product functions

def product_list():
    Book,Sheet = load()
    products = []
    for row in Sheet.iter_rows(values_only=True):
        products.append(list(row))

    for product in products:
        print(f"{product[0]}   |   {product[1]}   |   {product[2]}")
        print("-------------------------------------------------------------------")

    ans = input('Continue? Y >>> ').lower()
    startScreen()


def product_add():
    Book,Sheet = load()
    name = input("Product name >>> ")
    qty = input("Current quantity >>> ")
    if qty.isdigit() == False:
        print("Invalid quantity")
        startScreen()

    max = Sheet.max_row + 1

    NextID = Book.active.cell(row=1, column=999).value

    Sheet.cell(row=max, column=1).value = NextID
    Sheet.cell(row=max, column=2).value = name
    Sheet.cell(row=max, column=3).value = int(qty)

    Book.active.cell(row=1, column=999).value = Book.active.cell(row=1, column=999).value + 1

    Book.save('Data.xlsx')

    print(name + " was added with a quantity of " + str(qty) + " and an ID of " + str(NextID))

    ans = input('Add more? Y/N >>> ').lower()
    if ans == "y":
        product_add()
    else:
        print('Unknown command')
        startScreen()


def product_remove(id):
    Book,Sheet = load()

    name, qty, i = finder(id)

    if finder(id) == (None, None, None):
        print("Product not found")
        startScreen()

    Sheet.delete_rows(i)
    Book.save('Data.xlsx')
    Choice = input(name + " was removed from the list, remove another one? >> ").lower()
    def yes_choice():
        id = input('Input ID of product >>> ')
        product_remove(id)
    def no_choice():
        clearscreen()
        startScreen()
    def default_choice():
        print("Unknown command")
        clearscreen()
        startScreen()
    switcher = {
        "y": yes_choice,
        "n": no_choice,
    }

    func = switcher.get(Choice, default_choice)

    func()


def clearscreen():
    i = 0
    while i < 20:
        i = i + 1
        print("""
        
        
        
        
        
        """)

def finder(id):

    if not id.isdigit():
        return None, None, None

    Book,Sheet = load()
    i = 0
    for row in Sheet:
        i = i + 1
        if Book.active.cell(row=i, column=1).value == int(id):
            name = Book.active.cell(row=i, column=2).value
            qty = Book.active.cell(row=i, column=3).value
            return name, qty, i
    return None, None, None


def start_gui():
    choice = input("Start GUI mode? Y/N >>> ").lower()

    if choice == "y":
        subprocess.run(['python', 'UI.py'])
    else:
        startScreen()
