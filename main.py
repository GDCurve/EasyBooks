import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import PySimpleGUI as sg
from func import startScreen
import os


# Check for file
if os.path.exists('Data.xlsx') == True:
    Book = load_workbook('Data.xlsx')
    print('Workbook Found!')
else:
    Book = Workbook()
    Top = [["ID", "NAME", "COUNT"]]
    for row in Top:
        Book.active.append(row)
    print("Created Workbook / Not found")

Sheet = Book.active

print(r"""   .----------------.  .----------------.  .----------------.  .----------------.  .----------------.  .----------------.  .----------------.  .----------------.  .----------------. 
| .--------------. || .--------------. || .--------------. || .--------------. || .--------------. || .--------------. || .--------------. || .--------------. || .--------------. |
| |  _________   | || |      __      | || |    _______   | || |  ____  ____  | || |   ______     | || |     ____     | || |     ____     | || |  ___  ____   | || |    _______   | |
| | |_   ___  |  | || |     /  \     | || |   /  ___  |  | || | |_  _||_  _| | || |  |_   _ \    | || |   .'    `.   | || |   .'    `.   | || | |_  ||_  _|  | || |   /  ___  |  | |
| |   | |_  \_|  | || |    / /\ \    | || |  |  (__ \_|  | || |   \ \  / /   | || |    | |_) |   | || |  /  .--.  \  | || |  /  .--.  \  | || |   | |_/ /    | || |  |  (__ \_|  | |
| |   |  _|  _   | || |   / ____ \   | || |   '.___`-.   | || |    \ \/ /    | || |    |  __'.   | || |  | |    | |  | || |  | |    | |  | || |   |  __'.    | || |   '.___`-.   | |
| |  _| |___/ |  | || | _/ /    \ \_ | || |  |`\____) |  | || |    _|  |_    | || |   _| |__) |  | || |  \  `--'  /  | || |  \  `--'  /  | || |  _| |  \ \_  | || |  |`\____) |  | |
| | |_________|  | || ||____|  |____|| || |  |_______.'  | || |   |______|   | || |  |_______/   | || |   `.____.'   | || |   `.____.'   | || | |____||____| | || |  |_______.'  | |
| |              | || |              | || |              | || |              | || |              | || |              | || |              | || |              | || |              | |
| '--------------' || '--------------' || '--------------' || '--------------' || '--------------' || '--------------' || '--------------' || '--------------' || '--------------' |
 '----------------'  '----------------'  '----------------'  '----------------'  '----------------'  '----------------'  '----------------'  '----------------'  '----------------'   """)

print("""Welcome, please choose an option:\n
    add | read | write | help | exit""")



startScreen()



# save
Book.save('Data.xlsx')

